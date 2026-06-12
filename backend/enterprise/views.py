import os
import json
import hashlib
import secrets
import base64
import uuid
from pathlib import Path
from datetime import timedelta

import jwt
from django.conf import settings
from django.contrib.auth import authenticate, get_user_model
from django.db import connections
from django.db.models import Avg, Count, Sum
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from fpdf import FPDF
from google.oauth2 import id_token
from google.auth.transport import requests

from ml_api.model_service import ModelService
from dwh_sync.models import DWHEventData

from .models import (
    ActivityLog,
    AnalyticsSnapshot,
    AuditLog,
    DashboardConfig,
    FaceCheckIn,
    FaceProfile,
    EnterpriseExportLog,
    Notification,
    PasswordReset,
    Profile,
    RefreshSession,
    Role,
)

User = get_user_model()


def _duration_to_seconds(value: str, fallback: int) -> int:
    if not value:
        return fallback
    raw = str(value).strip().lower()
    if len(raw) < 2:
        return fallback
    unit = raw[-1]
    number = raw[:-1]
    if not number.isdigit():
        return fallback
    amount = int(number)
    multiplier = {"s": 1, "m": 60, "h": 3600, "d": 86400}.get(unit)
    if not multiplier:
        return fallback
    return amount * multiplier


ACCESS_TTL_SECONDS = _duration_to_seconds(
    getattr(settings, "ENTERPRISE_ACCESS_TTL", "15m"),
    900,
)
REFRESH_TTL_SECONDS = _duration_to_seconds(
    getattr(settings, "ENTERPRISE_REFRESH_TTL", "7d"),
    7 * 86400,
)


def _hash_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _issue_tokens(user: User, request) -> tuple[str, str, Role]:
    profile = _ensure_profile(user)
    role = profile.role
    access_payload = {
        "sub": str(user.id),
        "role": role.slug,
        "permissions": role.permissions,
        "exp": timezone.now() + timedelta(seconds=ACCESS_TTL_SECONDS),
    }
    refresh_payload = {
        "sub": str(user.id),
        "exp": timezone.now() + timedelta(seconds=REFRESH_TTL_SECONDS),
    }

    access_token = jwt.encode(access_payload, settings.ENTERPRISE_JWT_SECRET, algorithm="HS256")
    refresh_token = jwt.encode(refresh_payload, settings.ENTERPRISE_REFRESH_SECRET, algorithm="HS256")

    RefreshSession.objects.create(
        user=user,
        token_hash=_hash_token(refresh_token),
        user_agent=request.META.get("HTTP_USER_AGENT", ""),
        ip_address=request.META.get("REMOTE_ADDR", ""),
        expires_at=timezone.now() + timedelta(seconds=REFRESH_TTL_SECONDS),
    )

    return access_token, refresh_token, role


def _ensure_profile(user: User) -> Profile:
    profile = Profile.objects.select_related("role").filter(user=user).first()
    if profile:
        return profile
    role = Role.objects.filter(slug="business").first() or Role.objects.first()
    if not role:
        role = Role.objects.create(name="Business Manager", slug="business", permissions=[])
    return Profile.objects.create(user=user, role=role)


def _serialize_user(user: User, profile: Profile) -> dict:
    return {
        "id": user.id,
        "firstName": user.first_name,
        "lastName": user.last_name,
        "email": user.email,
        "role": profile.role.slug,
        "roleName": profile.role.name,
        "title": profile.title,
        "department": profile.department,
        "avatarUrl": profile.avatar_url,
        "preferences": profile.preferences,
    }


def _get_request_json(request) -> dict:
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        payload = {}
    if not isinstance(payload, dict):
        return {}
    return payload


def _log_activity(user: User | None, action: str, entity: str = "", metadata: dict | None = None, request=None) -> None:
    ActivityLog.objects.create(
        user=user,
        action=action,
        entity=entity,
        metadata=metadata or {},
        ip_address=request.META.get("REMOTE_ADDR", "") if request else "",
        user_agent=request.META.get("HTTP_USER_AGENT", "") if request else "",
    )


def _get_auth_context(request):
    bearer = request.META.get("HTTP_AUTHORIZATION", "")
    token = ""
    if bearer.startswith("Bearer "):
        token = bearer[7:]
    if not token:
        token = request.COOKIES.get("accessToken", "")
    if not token:
        return None, JsonResponse({"error": "Unauthorized"}, status=401)

    try:
        payload = jwt.decode(token, settings.ENTERPRISE_JWT_SECRET, algorithms=["HS256"])
    except Exception:
        return None, JsonResponse({"error": "Unauthorized"}, status=401)

    user = User.objects.filter(id=payload.get("sub")).first()
    if not user:
        return None, JsonResponse({"error": "Unauthorized"}, status=401)

    profile = _ensure_profile(user)
    return {
        "user": user,
        "profile": profile,
        "role": profile.role,
        "permissions": profile.role.permissions or [],
    }, None


def _require_permission(context: dict, permission: str) -> bool:
    permissions = context.get("permissions", [])
    return permission in permissions


def _rows_from_cursor(cursor) -> list[dict]:
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


CITY_COORDS = {
    "tunis": (36.8065, 10.1815),
    "ariana": (36.8663, 10.1934),
    "ben arous": (36.7531, 10.2319),
    "manouba": (36.8078, 10.1011),
    "sousse": (35.8256, 10.6084),
    "sfax": (34.7406, 10.7603),
    "monastir": (35.7643, 10.8113),
    "mahdia": (35.5047, 11.0622),
    "nabeul": (36.4513, 10.7351),
    "hammamet": (36.4000, 10.6167),
    "bizerte": (37.2722, 9.8710),
    "kairouan": (35.6769, 10.1010),
    "gabes": (33.8815, 10.0982),
    "gafsa": (34.4311, 8.7757),
    "tozeur": (33.9197, 8.1335),
    "djerba": (33.8076, 10.8451),
    "houmt souk": (33.8758, 10.8578),
    "zarzis": (33.5039, 11.1127),
    "kebili": (33.7050, 8.9679),
}


def _normalize_city_key(value: str) -> str:
    return str(value or "").strip().lower()


def _get_city_coords(city: str):
    key = _normalize_city_key(city)
    if not key:
        return None
    return CITY_COORDS.get(key)


def _pick_first_text(row: dict, keys: list[str]) -> str:
    for key in keys:
        value = row.get(key)
        text = str(value or "").strip()
        if text and text.lower() != "nan":
            return text
    return ""


def _pick_first_float(row: dict, keys: list[str]) -> float | None:
    for key in keys:
        value = row.get(key)
        if value is None or value == "":
            continue
        try:
            return float(value)
        except (TypeError, ValueError):
            continue
    return None


def _face_storage_dir() -> Path:
    base = Path(settings.BASE_DIR) / "media" / "face_profiles"
    base.mkdir(parents=True, exist_ok=True)
    return base


def _decode_image_data(data_url: str) -> tuple[bytes | None, str | None]:
    if not data_url:
        return None, "Missing image"

    raw = data_url
    if "," in raw:
        raw = raw.split(",", 1)[1]

    try:
        return base64.b64decode(raw), None
    except Exception:
        return None, "Invalid base64 image"


def _save_face_image(data_url: str) -> str | None:
    image_bytes, error = _decode_image_data(data_url)
    if error or not image_bytes:
        return None

    storage = _face_storage_dir()
    filename = f"{uuid.uuid4().hex}.jpg"
    image_path = storage / filename

    try:
        image_path.write_bytes(image_bytes)
    except Exception:
        return None

    return str(image_path)


def _normalize_descriptor(value) -> list[float] | None:
    if not isinstance(value, list):
        return None
    cleaned: list[float] = []
    for item in value:
        try:
            cleaned.append(float(item))
        except (TypeError, ValueError):
            return None
    return cleaned if cleaned else None


def _descriptor_distance(a: list[float], b: list[float]) -> float:
    if len(a) != len(b):
        return float("inf")
    total = 0.0
    for x, y in zip(a, b, strict=False):
        diff = x - y
        total += diff * diff
    return total ** 0.5


def _get_face_api_threshold() -> float:
    raw = os.getenv("FACE_API_THRESHOLD", "0.6")
    try:
        return float(raw)
    except (TypeError, ValueError):
        return 0.6


def _fetch_dwh_kpis() -> dict:
    with connections["dwh"].cursor() as cursor:
        cursor.execute("SELECT COUNT(*) FROM dim_reservation")
        reservations = cursor.fetchone()[0] or 0

        cursor.execute("SELECT COALESCE(SUM(event_budget), 0), COALESCE(AVG(event_budget), 0) FROM dim_event")
        total_budget, avg_budget = cursor.fetchone()

        cursor.execute("SELECT COALESCE(AVG(visitors), 0) FROM dim_visitors")
        avg_visitors = cursor.fetchone()[0] or 0

    return {
        "reservations": int(reservations or 0),
        "total_budget": float(total_budget or 0),
        "avg_budget": float(avg_budget or 0),
        "avg_visitors": float(avg_visitors or 0),
    }


def _fetch_marketing_trend() -> list[dict]:
    with connections["dwh"].cursor() as cursor:
        cursor.execute(
            """
            SELECT DATE_FORMAT(date_marketing, '%Y-%m') AS month,
                   SUM(marketing_spend) AS spend,
                   SUM(new_beneficiaries) AS new_beneficiaries
            FROM dim_marketing
            WHERE date_marketing IS NOT NULL
            GROUP BY month
            ORDER BY month
            """
        )
        rows = _rows_from_cursor(cursor)
    return [
        {
            "month": row.get("month") or "",
            "spend": float(row.get("spend") or 0),
            "new_beneficiaries": int(row.get("new_beneficiaries") or 0),
        }
        for row in rows
    ]


@csrf_exempt
def signup(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    payload = _get_request_json(request)
    first_name = str(payload.get("firstName", "")).strip()
    last_name = str(payload.get("lastName", "")).strip()
    email = str(payload.get("email", "")).strip().lower()
    password = str(payload.get("password", "")).strip()
    role_slug = str(payload.get("role", "business")).strip().lower()

    if not first_name or not last_name or not email or not password:
        return JsonResponse({"error": "Missing required fields"}, status=400)

    if User.objects.filter(email=email).exists():
        return JsonResponse({"error": "Email already in use"}, status=409)

    role = Role.objects.filter(slug=role_slug).first() or Role.objects.filter(slug="business").first()
    if not role:
        role = Role.objects.create(name="Business Manager", slug="business", permissions=[])

    user = User.objects.create_user(
        username=email,
        email=email,
        password=password,
        first_name=first_name,
        last_name=last_name,
    )
    profile = Profile.objects.create(user=user, role=role)

    access_token, refresh_token, role = _issue_tokens(user, request)
    _log_activity(user, "signup", "auth", {"email": email}, request)

    return JsonResponse(
        {
            "user": _serialize_user(user, profile),
            "accessToken": access_token,
            "refreshToken": refresh_token,
        }
    )


@csrf_exempt
def login(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    payload = _get_request_json(request)
    email = str(payload.get("email", "")).strip().lower()
    password = str(payload.get("password", "")).strip()
    if not email or not password:
        return JsonResponse({"error": "Missing credentials"}, status=400)

    user = authenticate(request, username=email, password=password)
    if not user:
        return JsonResponse({"error": "Invalid credentials"}, status=401)

    profile = _ensure_profile(user)
    user.last_login = timezone.now()
    user.save(update_fields=["last_login"])

    access_token, refresh_token, role = _issue_tokens(user, request)
    _log_activity(user, "login", "auth", {"email": email}, request)

    return JsonResponse(
        {
            "user": _serialize_user(user, profile),
            "accessToken": access_token,
            "refreshToken": refresh_token,
        }
    )


@csrf_exempt
def google_login(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    payload = _get_request_json(request)
    access_token_str = payload.get("access_token")
    if not access_token_str:
        return JsonResponse({"error": "Missing access token"}, status=400)

    try:
        import requests as py_requests
        # Verify the access token by calling Google's userinfo endpoint
        response = py_requests.get(
            "https://www.googleapis.com/oauth2/v3/userinfo",
            headers={"Authorization": f"Bearer {access_token_str}"}
        )
        
        if not response.ok:
            return JsonResponse({"error": "Invalid access token"}, status=401)
            
        idinfo = response.json()
        email = idinfo["email"]
        first_name = idinfo.get("given_name", "")
        last_name = idinfo.get("family_name", "")

        # Find or create user
        user, created = User.objects.get_or_create(
            email=email,
            defaults={
                "username": email,
                "first_name": first_name,
                "last_name": last_name,
            },
        )
        
        profile = _ensure_profile(user)
        access_token, refresh_token, role = _issue_tokens(user, request)
        
        return JsonResponse(
            {
                "user": _serialize_user(user, profile),
                "accessToken": access_token,
                "refreshToken": refresh_token,
            }
        )
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def refresh(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    payload = _get_request_json(request)
    refresh_token = payload.get("refreshToken") or request.COOKIES.get("refreshToken")
    if not refresh_token:
        return JsonResponse({"error": "Missing refresh token"}, status=400)

    try:
        decoded = jwt.decode(refresh_token, settings.ENTERPRISE_REFRESH_SECRET, algorithms=["HS256"])
    except Exception:
        return JsonResponse({"error": "Invalid refresh token"}, status=401)

    session = RefreshSession.objects.filter(
        user_id=decoded.get("sub"),
        token_hash=_hash_token(refresh_token),
        revoked_at__isnull=True,
    ).first()
    if not session or session.expires_at < timezone.now():
        return JsonResponse({"error": "Session expired"}, status=401)

    user = User.objects.filter(id=decoded.get("sub")).first()
    if not user:
        return JsonResponse({"error": "User not found"}, status=404)

    session.revoked_at = timezone.now()
    session.save(update_fields=["revoked_at"])

    profile = _ensure_profile(user)
    access_token, refresh_token, role = _issue_tokens(user, request)

    return JsonResponse(
        {
            "user": _serialize_user(user, profile),
            "accessToken": access_token,
            "refreshToken": refresh_token,
        }
    )


@csrf_exempt
def logout(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    payload = _get_request_json(request)
    refresh_token = payload.get("refreshToken") or request.COOKIES.get("refreshToken")
    if refresh_token:
        RefreshSession.objects.filter(token_hash=_hash_token(refresh_token), revoked_at__isnull=True).update(
            revoked_at=timezone.now()
        )

    context, _ = _get_auth_context(request)
    if context:
        _log_activity(context["user"], "logout", "auth", request=request)

    return JsonResponse({"message": "Logged out"})


@csrf_exempt
def forgot_password(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    payload = _get_request_json(request)
    email = str(payload.get("email", "")).strip().lower()
    if not email:
        return JsonResponse({"error": "Email is required"}, status=400)

    user = User.objects.filter(email=email).first()
    if not user:
        return JsonResponse({"message": "If the account exists, a reset email will be sent."})

    token = secrets.token_urlsafe(24)
    PasswordReset.objects.create(
        user=user,
        token_hash=_hash_token(token),
        expires_at=timezone.now() + timedelta(hours=1),
    )
    _log_activity(user, "password_reset_requested", "auth", {"email": email}, request)

    return JsonResponse({"message": "Password reset initiated.", "resetToken": token})


@csrf_exempt
def reset_password(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    payload = _get_request_json(request)
    token = str(payload.get("token", "")).strip()
    new_password = str(payload.get("newPassword", "")).strip()
    if not token or not new_password:
        return JsonResponse({"error": "Missing token or password"}, status=400)

    token_hash = _hash_token(token)
    record = PasswordReset.objects.filter(token_hash=token_hash, used_at__isnull=True).first()
    if not record or record.expires_at < timezone.now():
        return JsonResponse({"error": "Invalid or expired token"}, status=400)

    user = record.user
    user.set_password(new_password)
    user.save()

    record.used_at = timezone.now()
    record.save(update_fields=["used_at"])

    _log_activity(user, "password_reset", "auth", {"email": user.email}, request)

    return JsonResponse({"message": "Password updated."})


@csrf_exempt
def me(request):
    context, error = _get_auth_context(request)
    if error:
        return error

    if request.method == "GET":
        profile = context["profile"]
        return JsonResponse({"user": _serialize_user(context["user"], profile)})

    if request.method == "PUT":
        payload = _get_request_json(request)
        profile = context["profile"]
        user = context["user"]

        user.first_name = payload.get("firstName", user.first_name)
        user.last_name = payload.get("lastName", user.last_name)
        user.save(update_fields=["first_name", "last_name"])

        profile.title = payload.get("title", profile.title)
        profile.department = payload.get("department", profile.department)
        profile.avatar_url = payload.get("avatarUrl", profile.avatar_url)
        profile.preferences = payload.get("preferences", profile.preferences)
        profile.save()

        _log_activity(user, "profile_updated", "user", {"email": user.email}, request)
        return JsonResponse({"user": _serialize_user(user, profile)})

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
def ai_profile_suggestions(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    user = context["user"]
    profile = context["profile"]
    role_slug = profile.role.slug

    # Simple logic-based AI simulation that can be extended with real LLMs
    recommendations = {
        "ceo": {
            "persona": "Architecte de Croissance Stratégique",
            "focus": [
                "Surveillance de la marge brute globale via le modèle Price Regression",
                "Analyse des opportunités d'expansion sur les marchés à haut budget",
                "Optimisation du ROI marketing inter-canaux"
            ],
            "recommended_models": ["price_regression", "demand_forecast"]
        },
        "marketing": {
            "persona": "Catalyseur d'Engagement Client",
            "focus": [
                "Ciblage prédictif basé sur le comportement de réservation",
                "Optimisation du coût d'acquisition client (CAC)",
                "Analyse de sentiment sur les retours d'événements"
            ],
            "recommended_models": ["demand_forecast"]
        },
        "business": {
            "persona": "Stratège d'Efficacité Opérationnelle",
            "focus": [
                "Gestion dynamique des stocks et ressources",
                "Optimisation des prix de vente pour maximiser le taux d'occupation",
                "Réduction des coûts opérationnels via l'analyse de flux"
            ],
            "recommended_models": ["price_regression"]
        }
    }

    data = recommendations.get(role_slug, recommendations["business"])
    
    return JsonResponse({
        "role": role_slug,
        "persona": data["persona"],
        "recommendations": data["focus"],
        "recommendedModels": data["recommended_models"],
        "impactScore": 92 if role_slug == "ceo" else 88
    })


@csrf_exempt
def eagle_eye_alerts(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    # Simulated anomaly detection logic
    alerts = [
        {
            "id": "risk-1",
            "title": "Alerte de Risque Budgétaire",
            "detail": "Les prix des prestataires à Tunis ont augmenté de 12% ce mois-ci. Risque de dépassement sur 3 événements.",
            "severity": "high",
            "model": "Price Regression",
            "suggestion": "Réviser les marges pour les contrats Q4."
        },
        {
            "id": "risk-2",
            "title": "Anomalie de Demande",
            "detail": "Hausse inhabituelle des demandes de 'Mariage' à Sousse. +45% vs moyenne historique.",
            "severity": "medium",
            "model": "Demand Forecast",
            "suggestion": "Augmenter la capacité des prestataires locaux."
        }
    ]
    
    return JsonResponse({"alerts": alerts})


@csrf_exempt
def top_providers(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    # Publicly accessible AI-driven provider suggestions
    providers = [
        {
            "name": "Luxury Palace Tunis",
            "type": "Hôtel / Salle",
            "rating": 4.9,
            "offer": "-15% sur les mariages d'été",
            "image": "https://images.unsplash.com/photo-1519167758481-83f550bb49b3?w=800&q=80"
        },
        {
            "name": "Elite Catering Services",
            "type": "Traiteur",
            "rating": 4.8,
            "offer": "Cocktail offert pour > 200 pers.",
            "image": "https://images.unsplash.com/photo-1555244162-803834f70033?w=800&q=80"
        },
        {
            "name": "Magic Moments Photography",
            "type": "Photographe",
            "rating": 5.0,
            "offer": "Album Premium inclus",
            "image": "https://images.unsplash.com/photo-1537633552985-df8429e8048b?w=800&q=80"
        }
    ]
    
    return JsonResponse({
        "providers": providers,
        "message": "Ces offres sont exclusivement réservées aux membres EventZella."
    })


@csrf_exempt
def recommend_providers_by_budget(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        data = json.loads(request.body)
        budget = float(data.get("budget", 0))
        event_type = data.get("event_type", "Wedding")
        city = data.get("city", "Tunis")
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid input data"}, status=400)

    try:
        # Call the actual ML model service
        result = ModelService.predict("provider_budget_model", {
            "budget": budget,
            "event_type": event_type,
            "city": city,
            "top_k": 6
        })
        
        # Format the response for the frontend
        providers = []
        for rec in result.get("recommendations", []):
            providers.append({
                "name": rec.get("provider", "Elite Partner"),
                "type": rec.get("specialty", "Event Specialist"),
                "rating": 4.8, # Static for now as models don't always have rating
                "offer": f"Exclusive offer for {event_type}",
                "image": "https://images.unsplash.com/photo-1519167758481-83f550bb49b3?w=800&q=80",
                "price_indicator": rec.get("avg_price", 0)
            })

        return JsonResponse({
            "providers": providers,
            "message": "For these providers, you can book them at this price only with us EventZella."
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def providers_map(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    raw_limit = request.GET.get("limit", "800")
    try:
        limit = max(1, min(int(raw_limit), 2000))
    except (TypeError, ValueError):
        limit = 800

    providers: list[dict] = []
    source = "dwh"

    try:
        with connections["dwh"].cursor() as cursor:
            cursor.execute("SELECT * FROM dim_provider LIMIT %s", [limit])
            rows = _rows_from_cursor(cursor)
    except Exception:
        rows = []

    if rows:
        for row in rows:
            name = _pick_first_text(row, ["provider_name", "name", "provider", "provider_label", "company_name"])
            city = _pick_first_text(row, ["provider_city", "city", "ville", "location"])
            if not name or not city:
                continue

            lat = _pick_first_float(row, ["latitude", "lat", "provider_lat", "provider_latitude"])
            lng = _pick_first_float(row, ["longitude", "lng", "lon", "provider_lon", "provider_longitude"])
            coords = _get_city_coords(city)

            providers.append(
                {
                    "name": name,
                    "city": city,
                    "lat": lat if lat is not None else (coords[0] if coords else None),
                    "lng": lng if lng is not None else (coords[1] if coords else None),
                }
            )

    if not providers:
        source = "model"
        ModelService._ensure_dynamic_registry_data()
        for rec in ModelService._provider_records:
            city = str(rec.get("city", "")).strip()
            if not city:
                continue
            coords = _get_city_coords(city)
            if not coords:
                continue
            providers.append(
                {
                    "name": str(rec.get("provider") or "Unknown Provider"),
                    "city": city,
                    "lat": coords[0],
                    "lng": coords[1],
                }
            )

    return JsonResponse({"providers": providers, "source": source})


@csrf_exempt
def face_enroll(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    payload = _get_request_json(request)
    display_name = str(payload.get("name", "")).strip()
    image_data = payload.get("image")
    descriptor_raw = payload.get("descriptor")
    consent = bool(payload.get("consent", False))

    if not display_name:
        return JsonResponse({"error": "Field 'name' is required."}, status=400)
    if not consent:
        return JsonResponse({"error": "Consent is required."}, status=400)

    descriptor = _normalize_descriptor(descriptor_raw)
    if not descriptor:
        return JsonResponse({"error": "Field 'descriptor' is required."}, status=400)

    image_path = ""
    if image_data:
        image_path = _save_face_image(image_data) or ""

    profile = FaceProfile.objects.create(
        display_name=display_name,
        image_path=image_path,
        descriptor=descriptor,
        consent=True,
        consent_at=timezone.now(),
    )

    _log_activity(None, "face_enroll", entity=display_name, metadata={"profile_id": profile.id}, request=request)

    return JsonResponse({"status": "ok", "profile_id": profile.id, "name": profile.display_name})


@csrf_exempt
def face_verify(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    payload = _get_request_json(request)
    descriptor_raw = payload.get("descriptor")
    descriptor = _normalize_descriptor(descriptor_raw)
    if not descriptor:
        return JsonResponse({"error": "Field 'descriptor' is required."}, status=400)

    threshold = _get_face_api_threshold()
    best_profile = None
    best_distance = float("inf")

    for profile in FaceProfile.objects.all():
        stored = _normalize_descriptor(profile.descriptor)
        if not stored:
            continue
        distance = _descriptor_distance(descriptor, stored)
        if distance < best_distance:
            best_distance = distance
            best_profile = profile

    matched = bool(best_profile and best_distance <= threshold)
    matched_name = best_profile.display_name if matched else ""

    FaceCheckIn.objects.create(
        profile=best_profile if matched else None,
        matched_name=matched_name,
        confidence=None if best_distance == float("inf") else float(best_distance),
        success=matched,
        ip_address=request.META.get("REMOTE_ADDR", ""),
    )

    _log_activity(
        None,
        "face_checkin",
        entity=matched_name or "unknown",
        metadata={"distance": None if best_distance == float("inf") else float(best_distance), "matched": matched},
        request=request,
    )

    return JsonResponse(
        {
            "matched": matched,
            "name": matched_name,
            "profile_id": best_profile.id if matched else None,
            "distance": None if best_distance == float("inf") else round(float(best_distance), 4),
            "confidence": None if best_distance == float("inf") else round(float(best_distance), 4),
            "threshold": threshold,
        }
    )


@csrf_exempt
def roles(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    roles_list = list(Role.objects.all().values("name", "slug", "description", "permissions"))
    return JsonResponse({"roles": roles_list})


def _build_dashboard(role_slug: str) -> dict:
    role_profiles = {
        "ceo": {
            "name": "CEO",
            "focus": "Enterprise performance and strategic outlook",
            "kpis": [
                {"label": "Global Revenue", "value": 184.6, "unit": "M", "delta": 6.4},
                {"label": "ROI", "value": 28.9, "unit": "%", "delta": 1.8},
                {"label": "Conversion Rate", "value": 14.2, "unit": "%", "delta": 0.6},
                {"label": "Pipeline Health", "value": 92, "unit": "%", "delta": 2.4},
            ],
        },
        "quality": {
            "name": "Quality Manager",
            "focus": "Customer experience, risk, and compliance",
            "kpis": [
                {"label": "Satisfaction Index", "value": 91.4, "unit": "%", "delta": 1.2},
                {"label": "Complaint Rate", "value": 2.8, "unit": "%", "delta": -0.4},
                {"label": "SLA Adherence", "value": 96.1, "unit": "%", "delta": 0.8},
                {"label": "Risk Alerts", "value": 7, "unit": "", "delta": -1.0},
            ],
        },
        "business": {
            "name": "Business Manager",
            "focus": "Operational performance and growth",
            "kpis": [
                {"label": "Monthly Revenue", "value": 12.7, "unit": "M", "delta": 4.9},
                {"label": "Reservations", "value": 1284, "unit": "", "delta": 3.1},
                {"label": "Avg Order Value", "value": 987, "unit": "$", "delta": 2.4},
                {"label": "Utilization", "value": 83, "unit": "%", "delta": 1.7},
            ],
        },
        "marketing": {
            "name": "Marketing Manager",
            "focus": "Campaign performance and acquisition",
            "kpis": [
                {"label": "Marketing ROI", "value": 3.6, "unit": "x", "delta": 0.3},
                {"label": "CAC", "value": 42, "unit": "$", "delta": -3.2},
                {"label": "MQL to SQL", "value": 36.8, "unit": "%", "delta": 1.1},
                {"label": "Pipeline Influence", "value": 48.2, "unit": "%", "delta": 2.2},
            ],
        },
    }

    chart_templates = {
        "revenueTrend": [
            {"month": "Jan", "value": 9.4},
            {"month": "Feb", "value": 10.2},
            {"month": "Mar", "value": 11.1},
            {"month": "Apr", "value": 12.3},
            {"month": "May", "value": 12.7},
            {"month": "Jun", "value": 13.2},
        ],
        "satisfactionTrend": [
            {"month": "Jan", "value": 89},
            {"month": "Feb", "value": 90},
            {"month": "Mar", "value": 91},
            {"month": "Apr", "value": 92},
            {"month": "May", "value": 91.4},
            {"month": "Jun", "value": 93},
        ],
        "campaignPerformance": [
            {"name": "Brand", "value": 32},
            {"name": "Performance", "value": 28},
            {"name": "Partners", "value": 19},
            {"name": "Lifecycle", "value": 21},
        ],
        "qualityMix": [
            {"name": "Service Delay", "value": 34},
            {"name": "Product Quality", "value": 27},
            {"name": "Communication", "value": 22},
            {"name": "Logistics", "value": 17},
        ],
        "businessMix": [
            {"name": "Corporate", "value": 38},
            {"name": "Wedding", "value": 26},
            {"name": "Private", "value": 21},
            {"name": "Public", "value": 15},
        ],
        "marketingMix": [
            {"name": "Paid Search", "value": 31},
            {"name": "Social", "value": 27},
            {"name": "Partners", "value": 20},
            {"name": "Lifecycle", "value": 22},
        ],
        "funnel": [
            {"stage": "Awareness", "value": 9800},
            {"stage": "Engaged", "value": 5200},
            {"stage": "Qualified", "value": 2100},
            {"stage": "Converted", "value": 720},
        ],
    }

    recommendations_by_role = {
        "ceo": [
            {
                "title": "Rebalance premium portfolio",
                "impact": "High",
                "detail": "Shift 8% of high-margin capacity to corporate events in Q3.",
            },
            {
                "title": "Strengthen retention",
                "impact": "Medium",
                "detail": "Launch executive briefings for top 20 accounts.",
            },
        ],
        "quality": [
            {
                "title": "Escalate SLA monitoring",
                "impact": "High",
                "detail": "Introduce daily checks for high-risk vendors in the north region.",
            },
            {
                "title": "Feedback loop automation",
                "impact": "Medium",
                "detail": "Add automated remediation tasks for delayed service incidents.",
            },
        ],
        "business": [
            {
                "title": "Optimize resource allocation",
                "impact": "High",
                "detail": "Reassign crews to high-demand city clusters in the next cycle.",
            },
            {
                "title": "Margin guardrails",
                "impact": "Medium",
                "detail": "Apply dynamic pricing on premium segments above 180 guests.",
            },
        ],
        "marketing": [
            {
                "title": "Retarget high-fit segments",
                "impact": "High",
                "detail": "Increase paid social budget for top-performing wedding segments.",
            },
            {
                "title": "Lifecycle nurture",
                "impact": "Medium",
                "detail": "Launch win-back journeys for leads inactive for 30 days.",
            },
        ],
    }

    alerts_by_role = {
        "ceo": [
            {"title": "Portfolio risk review", "detail": "Two enterprise accounts approaching SLA threshold.", "type": "warning"},
            {"title": "Forecast variance", "detail": "Marketing lead quality improved 6% week over week.", "type": "info"},
        ],
        "quality": [
            {"title": "Quality escalation", "detail": "Complaint spikes detected in the logistics workflow.", "type": "warning"},
            {"title": "Risk watchlist", "detail": "Three events flagged for proactive QA review.", "type": "info"},
        ],
        "business": [
            {"title": "Capacity pressure", "detail": "Weekend utilization projected above 90%.", "type": "warning"},
            {"title": "Revenue momentum", "detail": "Corporate segment revenue up 8% month over month.", "type": "info"},
        ],
        "marketing": [
            {"title": "Campaign efficiency", "detail": "Paid search CPA decreased 12% in the last sprint.", "type": "info"},
            {"title": "Audience saturation", "detail": "Brand reach plateaued in two priority cities.", "type": "warning"},
        ],
    }

    role_charts = {
        "ceo": {"revenueTrend": chart_templates["revenueTrend"], "campaignPerformance": chart_templates["campaignPerformance"]},
        "quality": {"revenueTrend": chart_templates["satisfactionTrend"], "campaignPerformance": chart_templates["qualityMix"]},
        "business": {"revenueTrend": chart_templates["revenueTrend"], "campaignPerformance": chart_templates["businessMix"]},
        "marketing": {"revenueTrend": chart_templates["revenueTrend"], "campaignPerformance": chart_templates["marketingMix"]},
    }

    role_key = role_slug if role_slug in role_profiles else "ceo"
    profile = role_profiles[role_key]

    # Fetch the actual Role from DB to get the PowerBI Embed URL
    from enterprise.models import Role
    db_role = None
    try:
        db_role = Role.objects.get(slug=role_key)
        powerbi_url = db_role.powerbi_embed_url
    except Role.DoesNotExist:
        powerbi_url = None

    predictions = {}
    insights = []
    try:
        sample_inputs = {item["key"]: item.get("sample_input", {}) for item in ModelService.list_models()}

        if role_key == "ceo":
            provider_input = sample_inputs.get("provider_budget_model", {"budget": 12000, "event_type": "Corporate Event", "city": "Tunis", "top_k": 5})
            predictions["providers"] = ModelService.predict("provider_budget_model", provider_input)
            insights.append({"title": "Strategic provider coverage", "detail": "Provider recommendations align with demand bands.", "severity": "info"})

        if role_key == "quality":
            complaint_input = sample_inputs.get("complaint_risk_model", {"complaint_text": "Delayed setup and missing equipment caused disruption."})
            predictions["complaints"] = ModelService.predict("complaint_risk_model", complaint_input)
            cancellation_input = sample_inputs.get("cancellation_rate_model", {"event_type": "Wedding", "budget": 15000, "final_price": 16250})
            predictions["cancellations"] = ModelService.predict("cancellation_rate_model", cancellation_input)
            insights.append({"title": "Risk mitigation focus", "detail": "Complaint risk score drives proactive escalation alerts.", "severity": "warning"})

        if role_key == "business":
            price_input = sample_inputs.get("xgboost_price_regression", {"guests": 180, "event_type": "Corporate Event", "city": "Tunis", "budget": 18000})
            predictions["pricing"] = ModelService.predict("xgboost_price_regression", price_input)
            cluster_input = sample_inputs.get("kmeans_clustering", {"guests": 120, "price": 8000})
            predictions["segments"] = ModelService.predict("kmeans_clustering", cluster_input)
            insights.append({"title": "Revenue optimization", "detail": "Dynamic pricing remains above target margin.", "severity": "info"})

        if role_key == "marketing":
            forecast_input = sample_inputs.get("demand_forecast_ai", {"city": "Tunis", "event_type": "Wedding", "forecast_horizon": 6})
            predictions["forecast"] = ModelService.predict("demand_forecast_ai", forecast_input)
            svd_input = sample_inputs.get("svd_collaborative_filter", {"top_k": 3})
            predictions["recommendations"] = ModelService.predict("svd_collaborative_filter", svd_input)
            insights.append({"title": "Campaign timing", "detail": "Forecasted demand supports Q3 reallocation.", "severity": "success"})
    except Exception:
        insights = [
            {
                "title": "ML data unavailable",
                "detail": "Live model outputs could not be loaded. Using baseline metrics.",
                "severity": "warning",
            }
        ]

    dwh_kpis = None
    try:
        dwh_kpis = _fetch_dwh_kpis()
    except Exception:
        dwh_kpis = None

    marketing_trend = []
    if role_key == "marketing":
        try:
            marketing_trend = _fetch_marketing_trend()
        except Exception:
            marketing_trend = []

    dynamic_kpis = [
        {"label": "Total Reservations (DWH)", "value": dwh_kpis["reservations"] if dwh_kpis else 0, "unit": "", "delta": 0},
        {"label": "Total Budget (DWH)", "value": round(dwh_kpis["total_budget"], 0) if dwh_kpis else 0, "unit": "TND", "delta": 0},
        {"label": "Avg Budget (DWH)", "value": round(dwh_kpis["avg_budget"], 0) if dwh_kpis else 0, "unit": "TND", "delta": 0},
        {"label": "Avg Visitors (DWH)", "value": round(dwh_kpis["avg_visitors"], 1) if dwh_kpis else 0, "unit": "", "delta": 0},
    ]

    return {
        "role": role_key,
        "roleName": profile["name"],
        "focus": profile["focus"],
        "powerbiEmbedUrl": powerbi_url,
        "allowedModels": db_role.allowed_ml_models if db_role else [],
        "kpis": dynamic_kpis if dwh_kpis and dwh_kpis["reservations"] > 0 else profile["kpis"],
        "charts": {
            "revenueTrend": role_charts[role_key]["revenueTrend"],
            "satisfactionTrend": chart_templates["satisfactionTrend"],
            "campaignPerformance": role_charts[role_key]["campaignPerformance"],
            "funnel": chart_templates["funnel"],
            "forecast": predictions.get("forecast", {}).get("forecast_points", []),
            "marketingTrend": marketing_trend,
        },
        "insights": insights,
        "predictions": predictions,
        "alerts": alerts_by_role.get(role_key, alerts_by_role["ceo"]),
        "recommendations": recommendations_by_role.get(role_key, recommendations_by_role["ceo"]),
        "lastUpdated": timezone.now().isoformat(),
    }


@csrf_exempt
def dashboard(request, role: str):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    requested = str(role or "").lower()
    profile_role = context["profile"].role.slug
    permissions = context["permissions"]
    role_permission_map = {
        "ceo": "view_all_dashboards",
        "quality": "view_quality_dashboard",
        "business": "view_business_dashboard",
        "marketing": "view_marketing_dashboard",
    }

    allowed = requested == profile_role or "view_all_dashboards" in permissions
    if not allowed and role_permission_map.get(requested) in permissions:
        allowed = True

    if not allowed:
        return JsonResponse({"error": "Forbidden"}, status=403)

    payload = _build_dashboard(requested)
    _log_activity(context["user"], "view_dashboard", requested, {"role": requested}, request)

    return JsonResponse(payload)


@csrf_exempt
def notifications_list(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    items = Notification.objects.filter(user=context["user"]).order_by("-created_at")[:20]
    data = [
        {
            "_id": item.id,
            "title": item.title,
            "message": item.message,
            "type": item.type,
            "readAt": item.read_at.isoformat() if item.read_at else None,
            "createdAt": item.created_at.isoformat(),
        }
        for item in items
    ]
    return JsonResponse({"notifications": data})


@csrf_exempt
def notification_read(request, notification_id: int):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    Notification.objects.filter(id=notification_id, user=context["user"]).update(read_at=timezone.now())
    return JsonResponse({"status": "ok"})


@csrf_exempt
def activity_list(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    logs = ActivityLog.objects.filter(user=context["user"]).order_by("-created_at")[:25]
    data = [
        {
            "_id": item.id,
            "action": item.action,
            "entity": item.entity,
            "createdAt": item.created_at.isoformat(),
        }
        for item in logs
    ]
    return JsonResponse({"logs": data})


@csrf_exempt
def analytics_history(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    role_slug = context["profile"].role.slug
    snapshots = AnalyticsSnapshot.objects.filter(role=role_slug).order_by("-generated_at")[:12]
    if not snapshots:
        return JsonResponse(
            {
                "history": [
                    {"month": "Jan", "value": 82},
                    {"month": "Feb", "value": 85},
                    {"month": "Mar", "value": 87},
                    {"month": "Apr", "value": 89},
                    {"month": "May", "value": 90},
                    {"month": "Jun", "value": 92},
                ]
            }
        )

    history = [
        {
            "month": snapshot.generated_at.strftime("%b"),
            "value": snapshot.metrics.get("value", 0) if snapshot.metrics else 0,
        }
        for snapshot in snapshots
    ]
    return JsonResponse({"history": history})


@csrf_exempt
def dwh_summary(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    try:
        dwh_kpis = _fetch_dwh_kpis()
        with connections["dwh"].cursor() as cursor:
            cursor.execute(
                """
                SELECT DATE_FORMAT(reservation_date, '%Y-%m') AS month,
                       COUNT(*) AS events
                FROM dim_reservation
                WHERE reservation_date IS NOT NULL
                GROUP BY month
                ORDER BY month
                """
            )
            monthly_events = _rows_from_cursor(cursor)

            cursor.execute(
                """
                SELECT DATE_FORMAT(event_date, '%Y-%m') AS month,
                       SUM(event_budget) AS budget
                FROM dim_event
                WHERE event_date IS NOT NULL
                GROUP BY month
                ORDER BY month
                """
            )
            monthly_budget = _rows_from_cursor(cursor)

            cursor.execute(
                """
                SELECT DATE_FORMAT(reservation_date, '%Y-%m') AS month,
                       SUM(CASE WHEN status = 'Confirmed' THEN 1 ELSE 0 END) AS confirmed,
                       SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) AS pending,
                       SUM(CASE WHEN status = 'Cancelled' THEN 1 ELSE 0 END) AS cancelled,
                       SUM(CASE WHEN status IS NULL OR status NOT IN ('Confirmed', 'Pending', 'Cancelled') THEN 1 ELSE 0 END) AS other
                FROM dim_reservation
                WHERE reservation_date IS NOT NULL
                GROUP BY month
                ORDER BY month
                """
            )
            reservation_status = _rows_from_cursor(cursor)

            cursor.execute(
                """
                SELECT DATE_FORMAT(date_visitor_m, '%Y-%m') AS month,
                       SUM(visitors) AS visitors,
                       SUM(reservations) AS reservations
                FROM dim_visitors
                WHERE date_visitor_m IS NOT NULL
                GROUP BY month
                ORDER BY month
                """
            )
            visitors = _rows_from_cursor(cursor)

            cursor.execute(
                """
                SELECT event_type AS name,
                       COUNT(*) AS events,
                       SUM(event_budget) AS budget
                FROM dim_event
                GROUP BY event_type
                ORDER BY events DESC
                LIMIT 6
                """
            )
            event_types = _rows_from_cursor(cursor)

            cursor.execute(
                """
                SELECT provider_city AS name,
                       COUNT(*) AS providers
                FROM dim_provider
                GROUP BY provider_city
                ORDER BY providers DESC
                LIMIT 6
                """
            )
            top_cities = _rows_from_cursor(cursor)

            cursor.execute(
                """
                SELECT DATE_FORMAT(date_marketing, '%Y-%m') AS month,
                       SUM(marketing_spend) AS spend,
                       SUM(new_beneficiaries) AS new_beneficiaries
                FROM dim_marketing
                WHERE date_marketing IS NOT NULL
                GROUP BY month
                ORDER BY month
                """
            )
            marketing = _rows_from_cursor(cursor)
    except Exception:
        return JsonResponse({"error": "DWH unavailable"}, status=500)

    monthly_index = {}
    for row in monthly_events:
        month = row.get("month") or ""
        monthly_index.setdefault(month, {"month": month, "events": 0, "budget": 0})
        monthly_index[month]["events"] = int(row.get("events") or 0)
    for row in monthly_budget:
        month = row.get("month") or ""
        monthly_index.setdefault(month, {"month": month, "events": 0, "budget": 0})
        monthly_index[month]["budget"] = float(row.get("budget") or 0)

    response = {
        "kpis": {
            "events": dwh_kpis.get("reservations", 0),
            "total_budget": float(dwh_kpis.get("total_budget", 0)),
            "avg_budget": float(dwh_kpis.get("avg_budget", 0)),
            "avg_guests": float(dwh_kpis.get("avg_visitors", 0)),
        },
        "monthly": sorted(monthly_index.values(), key=lambda item: item["month"]),
        "reservation_status": [
            {
                "month": row.get("month") or "",
                "confirmed": int(row.get("confirmed") or 0),
                "pending": int(row.get("pending") or 0),
                "cancelled": int(row.get("cancelled") or 0),
                "other": int(row.get("other") or 0),
            }
            for row in reservation_status
        ],
        "visitors": [
            {
                "month": row.get("month") or "",
                "visitors": int(row.get("visitors") or 0),
                "reservations": int(row.get("reservations") or 0),
            }
            for row in visitors
        ],
        "marketing": [
            {
                "month": row.get("month") or "",
                "spend": float(row.get("spend") or 0),
                "new_beneficiaries": int(row.get("new_beneficiaries") or 0),
            }
            for row in marketing
        ],
        "top_cities": [
            {
                "name": row.get("name") or "Unknown",
                "providers": int(row.get("providers") or 0),
            }
            for row in top_cities
        ],
        "event_types": [
            {
                "name": row.get("name") or "Unknown",
                "events": int(row.get("events") or 0),
                "budget": float(row.get("budget") or 0),
            }
            for row in event_types
        ],
    }

    return JsonResponse(response)


@csrf_exempt
def search(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    term = str(request.GET.get("q", "")).strip()
    if not term:
        return JsonResponse({"results": []})

    notifications = Notification.objects.filter(user=context["user"], title__icontains=term)[:10]
    activities = ActivityLog.objects.filter(user=context["user"], action__icontains=term)[:10]

    results = [
        {
            "type": "notification",
            "title": item.title,
            "description": item.message,
            "timestamp": item.created_at.isoformat(),
        }
        for item in notifications
    ]
    results.extend(
        {
            "type": "activity",
            "title": item.action,
            "description": item.entity,
            "timestamp": item.created_at.isoformat(),
        }
        for item in activities
    )

    return JsonResponse({"results": results})


@csrf_exempt
def admin_users(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    if not _require_permission(context, "manage_users"):
        return JsonResponse({"error": "Forbidden"}, status=403)

    users = User.objects.all().order_by("email")
    response = []
    for user in users:
        profile = _ensure_profile(user)
        response.append(
            {
                "id": user.id,
                "firstName": user.first_name,
                "lastName": user.last_name,
                "email": user.email,
                "role": profile.role.slug,
                "title": profile.title,
                "department": profile.department,
                "status": "active",
                "lastLoginAt": user.last_login.isoformat() if user.last_login else None,
            }
        )

    return JsonResponse({"users": response})


@csrf_exempt
def admin_update_role(request, user_id: int):
    if request.method != "PUT":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    if not _require_permission(context, "manage_users"):
        return JsonResponse({"error": "Forbidden"}, status=403)

    payload = _get_request_json(request)
    role_slug = str(payload.get("role", "")).strip().lower()
    role = Role.objects.filter(slug=role_slug).first()
    if not role:
        return JsonResponse({"error": "Invalid role"}, status=400)

    user = User.objects.filter(id=user_id).first()
    if not user:
        return JsonResponse({"error": "User not found"}, status=404)

    profile = _ensure_profile(user)
    profile.role = role
    profile.save(update_fields=["role"])

    _log_activity(context["user"], "role_changed", "user", {"targetUser": user.email, "newRole": role.slug}, request)

    return JsonResponse({"user": {"id": user.id, "email": user.email, "role": role.slug}})


@csrf_exempt
def export_excel(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    if not _require_permission(context, "export_data"):
        return JsonResponse({"error": "Forbidden"}, status=403)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Executive KPIs"
    sheet.append(["Metric", "Value", "Delta"])
    sheet.append(["Revenue", 12.7, 4.9])
    sheet.append(["Conversion", 14.2, 0.6])
    sheet.append(["Satisfaction", 91.4, 1.2])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=eventzella-kpis.xlsx"
    workbook.save(response)

    EnterpriseExportLog.objects.create(user=context["user"], export_type="excel")
    return response


@csrf_exempt
def export_pdf(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    if not _require_permission(context, "export_data"):
        return JsonResponse({"error": "Forbidden"}, status=403)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=14)
    pdf.cell(0, 10, txt="Eventzella Executive KPI Summary", ln=1)
    pdf.set_font("Arial", size=11)
    pdf.cell(0, 8, txt="Revenue: 12.7 (Delta 4.9)", ln=1)
    pdf.cell(0, 8, txt="Conversion: 14.2 (Delta 0.6)", ln=1)
    pdf.cell(0, 8, txt="Satisfaction: 91.4 (Delta 1.2)", ln=1)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=eventzella-kpis.pdf"
    response.write(pdf.output(dest="S").encode("latin-1"))

    EnterpriseExportLog.objects.create(user=context["user"], export_type="pdf")
    return response


@csrf_exempt
def audit_list(request):
    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    context, error = _get_auth_context(request)
    if error:
        return error

    if not _require_permission(context, "view_activity"):
        return JsonResponse({"error": "Forbidden"}, status=403)

    records = AuditLog.objects.order_by("-created_at")[:50]
    payload = [
        {
            "id": record.id,
            "eventType": record.event_type,
            "payload": record.payload,
            "createdAt": record.created_at.isoformat(),
        }
        for record in records
    ]
    return JsonResponse({"records": payload})
